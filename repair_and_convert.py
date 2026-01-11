import pandas as pd
import os
import zipfile
import shutil

files = [
    "/Users/ziyuexu/Documents/trae_projects/paper1/BMJ.xlsx",
    "/Users/ziyuexu/Documents/trae_projects/paper1/JAMA.xlsx",
    "/Users/ziyuexu/Documents/trae_projects/paper1/Lancet.xlsx"
]

def try_repair(file_path):
    print(f"--- Repairing {os.path.basename(file_path)} ---")
    
    # 1. Check if it's a valid ZIP (XLSX is a ZIP)
    if zipfile.is_zipfile(file_path):
        print("Format: Valid ZIP (XLSX structure)")
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                # Check for workbook.xml
                if 'xl/workbook.xml' in z.namelist():
                    print("Structure: Found workbook.xml")
                else:
                    print("Structure: Missing workbook.xml! This file is corrupted.")
        except Exception as e:
            print(f"ZIP Error: {e}")
    else:
        print("Format: NOT a valid ZIP. It might be CSV or XLS.")
        # Try reading as CSV
        try:
            df = pd.read_csv(file_path)
            print("Success: It was a CSV file disguised as XLSX!")
            # Convert to real XLSX
            new_path = file_path.replace('.xlsx', '_repaired.xlsx')
            df.to_excel(new_path, index=False)
            print(f"Repaired file saved to: {os.path.basename(new_path)}")
            return new_path
        except:
            print("Failed to read as CSV.")
            
        # Try reading as XLS (old format)
        try:
            df = pd.read_excel(file_path, engine='xlrd')
            print("Success: It was an old XLS file!")
            new_path = file_path.replace('.xlsx', '_repaired.xlsx')
            df.to_excel(new_path, index=False)
            return new_path
        except:
            print("Failed to read as XLS.")

    # Try openpyxl read_only mode
    try:
        from openpyxl import load_workbook
        print("Attempting openpyxl read_only=True...")
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if wb.sheetnames:
            print(f"Success! Found sheets: {wb.sheetnames}")
            ws = wb[wb.sheetnames[0]]
            data = ws.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            
            new_path = file_path.replace('.xlsx', '_repaired.xlsx')
            df.to_excel(new_path, index=False)
            print(f"Repaired file saved to: {os.path.basename(new_path)}")
            return new_path
        else:
            print("read_only mode found no sheets.")
    except Exception as e:
        print(f"openpyxl read_only failed: {e}")

    return None

for f in files:
    if os.path.exists(f):
        try_repair(f)
    else:
        print(f"File not found: {f}")
