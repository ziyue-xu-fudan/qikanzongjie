#!/usr/bin/env python3
"""
NEJM文献DOI筛选工具 - Excel版本
筛选DOI号包含"NEJMoa"的文献，并将其转换为Excel格式
"""

import pandas as pd
from pathlib import Path
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def filter_excel_by_doi_pattern(input_file, doi_pattern="NEJMoa"):
    """
    读取Excel文件，筛选DOI包含指定模式的文献，并创建新的Excel文件
    
    Args:
        input_file: 输入Excel文件路径
        doi_pattern: DOI匹配模式，默认为"NEJMoa"
    
    Returns:
        输出的Excel文件路径
    """
    print(f"🔍 开始Excel DOI筛选，模式: {doi_pattern}")
    print(f"📁 输入文件: {input_file}")
    
    try:
        # 读取Excel文件
        df = pd.read_excel(input_file)
        print(f"📊 原始文献总数: {len(df)} 篇")
        print(f"📋 数据列: {', '.join(df.columns)}")
        
        # 找到DOI列
        doi_column = None
        possible_doi_columns = ['DOI', 'doi', 'Doi', 'DOI_Number', 'doi_number']
        
        for col in possible_doi_columns:
            if col in df.columns:
                doi_column = col
                break
        
        if doi_column is None:
            # 查找包含DOI的列
            doi_columns = [col for col in df.columns if 'doi' in col.lower()]
            if doi_columns:
                doi_column = doi_columns[0]
            else:
                print("⚠️  未找到DOI列")
                return None
        
        print(f"✅ 使用DOI列: {doi_column}")
        
        # 显示DOI样本
        print("\n📄 DOI样本:")
        sample_dois = df[doi_column].dropna().head(5)
        for i, doi in enumerate(sample_dois, 1):
            print(f"  {i}. {doi}")
        
        # 筛选包含指定模式的DOI
        print(f"\n🎯 正在筛选包含'{doi_pattern}'的DOI...")
        
        # 使用正则表达式进行不区分大小写的匹配
        pattern = re.compile(doi_pattern, re.IGNORECASE)
        filtered_df = df[df[doi_column].notna() & df[doi_column].astype(str).str.contains(pattern, na=False)].copy()
        
        print(f"✅ 筛选完成！")
        print(f"📉 被过滤掉的文献: {len(df) - len(filtered_df)} 篇")
        print(f"📈 幸存的文献: {len(filtered_df)} 篇")
        print(f"💯 存活率: {(len(filtered_df) / len(df) * 100):.1f}%")
        
        # 显示筛选后的DOI样本
        if len(filtered_df) > 0:
            print(f"\n📄 筛选后的DOI样本:")
            filtered_dois = filtered_df[doi_column].head(5)
            for i, doi in enumerate(filtered_dois, 1):
                print(f"  {i}. {doi}")
        
        return filtered_df
        
    except FileNotFoundError:
        print(f"❌ 文件未找到: {input_file}")
        return None
    except Exception as e:
        print(f"❌ 处理文件时出错: {e}")
        return None

def create_formatted_excel(filtered_df, input_file, doi_pattern="NEJMoa"):
    """
    创建格式化的Excel文件
    """
    if filtered_df is None or filtered_df.empty:
        print("⚠️  没有数据需要保存")
        return None
    
    try:
        # 生成输出文件名
        input_path = Path(input_file)
        base_name = input_path.stem
        
        # 如果文件名已经包含筛选标记，先移除
        if "_doi_" in base_name:
            base_name = base_name.split("_doi_")[0]
        
        output_name = f"{base_name}_doi_{doi_pattern}.xlsx"
        output_file = input_path.parent / output_name
        
        # 创建Excel writer对象
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 将数据写入Excel
            filtered_df.to_excel(writer, sheet_name='Filtered_Literature', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['Filtered_Literature']
            
            # 设置列宽和样式
            setup_excel_formatting(worksheet, filtered_df)
            
        print(f"💾 格式化Excel文件已保存: {output_file}")
        return str(output_file)
        
    except Exception as e:
        print(f"❌ 创建Excel文件时出错: {e}")
        return None

def setup_excel_formatting(worksheet, df):
    """设置Excel格式化和样式"""
    # 标题行样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 数据行样式
    data_font = Font(size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用标题行样式
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置行高
        worksheet.row_dimensions[1].height = 30
    
    # 应用数据行样式和设置列宽
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
        
        # 设置数据行高
        worksheet.row_dimensions[row_num].height = 45
    
    # 智能设置列宽
    column_widths = {
        'PMID': 12,
        'pmid': 12,
        'Title': 50,
        'title': 50,
        'Authors': 40,
        'authors': 40,
        'DOI': 30,
        'doi': 30,
        'Journal': 20,
        'journal': 20,
        'Publication Year': 12,
        'pub_year': 12,
        'author_count': 12
    }
    
    # 为每列设置合适的宽度
    for col_num, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        if col_name in column_widths:
            worksheet.column_dimensions[col_letter].width = column_widths[col_name]
        else:
            # 默认宽度，根据内容长度调整
            max_length = max(df[col_name].astype(str).str.len().max(), len(col_name)) + 2
            worksheet.column_dimensions[col_letter].width = min(max_length, 50)
    
    # 添加筛选功能
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'
    
    # 如果存在author_count列，添加条件格式
    author_count_col = None
    for col_num, col_name in enumerate(df.columns, 1):
        if 'author_count' in col_name.lower():
            author_count_col = get_column_letter(col_num)
            break
    
    if author_count_col:
        # 为作者数量大于20的行添加特殊背景色
        high_author_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        from openpyxl.formatting.rule import CellIsRule
        
        rule = CellIsRule(operator='greaterThan', formula=['20'], fill=high_author_fill)
        worksheet.conditional_formatting.add(f"{author_count_col}2:{author_count_col}{len(df)+1}", rule)

def generate_statistics(filtered_df, original_count, doi_pattern):
    """生成统计信息"""
    stats = {
        '原始文献数': original_count,
        '筛选后文献数': len(filtered_df),
        '过滤掉的文献': original_count - len(filtered_df),
        '存活率': f"{(len(filtered_df) / original_count * 100):.1f}%",
        'DOI模式': doi_pattern
    }
    
    # DOI统计
    doi_column = None
    for col in filtered_df.columns:
        if 'doi' in col.lower():
            doi_column = col
            break
    
    if doi_column and len(filtered_df) > 0:
        doi_counts = {}
        for doi in filtered_df[doi_column].dropna():
            if '/' in str(doi):
                prefix = str(doi).split('/')[0]
                doi_counts[prefix] = doi_counts.get(prefix, 0) + 1
        
        if doi_counts:
            stats['DOI前缀分布'] = dict(sorted(doi_counts.items(), key=lambda x: x[1], reverse=True)[:5])
    
    return stats

def main():
    """主函数"""
    print("🔍 NEJM文献DOI筛选工具 - Excel版本")
    print("=" * 60)
    print("专门筛选DOI包含特定模式的NEJM文献并生成格式化Excel")
    print("=" * 60)
    
    # 输入文件路径
    input_file = "/Users/ziyuexu/Documents/trae_projects/paper1/csv-TheNewEngl-set (1)_authors_ge5.xlsx"
    
    # DOI匹配模式
    doi_pattern = "NEJMoa"
    
    print(f"📁 输入文件: {input_file}")
    print(f"🎯 筛选模式: {doi_pattern}")
    print()
    
    # 执行筛选
    filtered_df = filter_excel_by_doi_pattern(input_file, doi_pattern)
    
    if filtered_df is not None:
        # 获取原始数据数量
        original_df = pd.read_excel(input_file)
        original_count = len(original_df)
        
        # 生成统计信息
        stats = generate_statistics(filtered_df, original_count, doi_pattern)
        
        print(f"\n📊 筛选统计:")
        for key, value in stats.items():
            if key == 'DOI前缀分布':
                print(f"  {key}:")
                for prefix, count in value.items():
                    print(f"    {prefix}: {count} 篇")
            else:
                print(f"  {key}: {value}")
        
        # 创建格式化的Excel文件
        output_file = create_formatted_excel(filtered_df, input_file, doi_pattern)
        
        if output_file:
            print(f"\n🎉 任务完成！")
            print(f"📄 格式化Excel文件: {output_file}")
            print(f"📋 最终文献数量: {len(filtered_df)} 篇")
            print(f"✨ 文件已添加专业格式化和样式")
        else:
            print("\n❌ Excel文件创建失败")
    else:
        print("\n❌ 筛选过程失败")

if __name__ == "__main__":
    main()