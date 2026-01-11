#!/usr/bin/env python3
"""
NEJM文献Excel转换工具
将筛选后的CSV文件转换为格式化的Excel文件
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import re

def convert_csv_to_excel(input_csv_file):
    """
    将NEJM CSV文件转换为Excel格式，并添加格式化和样式
    
    Args:
        input_csv_file: 输入CSV文件路径
    
    Returns:
        输出的Excel文件路径
    """
    print(f"📊 开始转换CSV到Excel...")
    print(f"📁 输入文件: {input_csv_file}")
    
    try:
        # 读取CSV文件
        df = pd.read_csv(input_csv_file)
        print(f"📋 读取到 {len(df)} 篇文献")
        
        # 获取文件名（不含扩展名）
        input_path = Path(input_csv_file)
        output_excel = input_path.parent / f"{input_path.stem}.xlsx"
        
        # 创建Excel writer对象
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # 将数据写入Excel
            df.to_excel(writer, sheet_name='NEJM文献', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['NEJM文献']
            
            # 设置列宽
            column_widths = {
                'A': 12,  # PMID
                'B': 50,  # Title
                'C': 60,  # Authors
                'D': 50,  # Citation
                'E': 25,  # First Author
                'F': 20,  # Journal/Book
                'G': 12,  # Publication Year
                'H': 15,  # Create Date
                'I': 15,  # PMCID
                'J': 15,  # NIHMS ID
                'K': 25,  # DOI
                'L': 12   # author_count
            }
            
            # 应用列宽
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # 设置标题行样式
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 创建样式
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 数据行样式
            data_font = Font(size=10)
            data_alignment = Alignment(vertical="top", wrap_text=True)
            
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 应用标题行样式
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # 设置行高
                worksheet.row_dimensions[1].height = 30
            
            # 应用数据行样式
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                
                # 设置数据行高
                worksheet.row_dimensions[row_num].height = 60
            
            # 添加筛选功能
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # 添加条件格式（高作者数量高亮）
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill
            
            # 为作者数量大于20的行添加绿色背景
            green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            
            # 获取作者数量列的字母
            author_count_col = None
            for col_num, col_name in enumerate(df.columns, 1):
                if 'author_count' in col_name.lower():
                    author_count_col = get_column_letter(col_num)
                    break
            
            if author_count_col:
                # 添加条件格式规则
                rule = CellIsRule(operator='greaterThan', formula=['20'], fill=green_fill)
                worksheet.conditional_formatting.add(f"{author_count_col}2:{author_count_col}{len(df)+1}", rule)
            
            # 冻结首行
            worksheet.freeze_panes = 'A2'
            
            # 添加数据验证（Publication Year列）
            from openpyxl.worksheet.datavalidation import DataValidation
            
            # 创建年份范围验证
            year_validation = DataValidation(type="whole", operator="between", formula1="2020", formula2="2030")
            year_validation.error = "请输入2020-2030之间的年份"
            year_validation.errorTitle = "无效年份"
            
            # 找到Publication Year列
            year_col = None
            for col_num, col_name in enumerate(df.columns, 1):
                if 'year' in col_name.lower() and 'publication' in col_name.lower():
                    year_col = get_column_letter(col_num)
                    break
            
            if year_col:
                worksheet.add_data_validation(year_validation)
                year_validation.add(f"{year_col}2:{year_col}{len(df)+1}")
            
            print(f"✅ Excel文件创建完成: {output_excel}")
            
            # 生成文件统计信息
            stats = generate_excel_stats(df)
            print(f"📊 文件统计信息:")
            for key, value in stats.items():
                print(f"  {key}: {value}")
            
            return str(output_excel)
            
    except Exception as e:
        print(f"❌ 转换过程出错: {e}")
        return None

def generate_excel_stats(df):
    """生成Excel文件的统计信息"""
    stats = {}
    
    # 基本统计
    stats['总文献数'] = len(df)
    
    # 作者数量统计
    if 'author_count' in df.columns:
        stats['平均作者数'] = f"{df['author_count'].mean():.1f}"
        stats['最多作者数'] = df['author_count'].max()
        stats['最少作者数'] = df['author_count'].min()
    
    # 年份统计
    year_cols = [col for col in df.columns if 'year' in col.lower() and 'publication' in col.lower()]
    if year_cols:
        year_col = year_cols[0]
        year_counts = df[year_col].value_counts().sort_index(ascending=False)
        if len(year_counts) > 0:
            stats['最新文献年份'] = year_counts.index[0]
            stats['最活跃年份'] = year_counts.idxmax()
    
    # 期刊统计
    journal_cols = [col for col in df.columns if 'journal' in col.lower()]
    if journal_cols:
        journal_col = journal_cols[0]
        unique_journals = df[journal_col].nunique()
        stats['期刊种类数'] = unique_journals
    
    return stats

def main():
    """主函数"""
    print("📚 NEJM文献Excel转换工具")
    print("=" * 50)
    
    # 输入文件路径
    input_csv = "/Users/ziyuexu/Documents/trae_projects/paper1/csv-TheNewEngl-set (1)_authors_ge5.csv"
    
    print(f"🔄 开始转换: {input_csv}")
    
    # 执行转换
    output_excel = convert_csv_to_excel(input_csv)
    
    if output_excel:
        print(f"\n🎉 转换成功！")
        print(f"📄 Excel文件: {output_excel}")
        print(f"💡 提示: Excel文件已添加格式化、筛选、条件格式等功能")
        print(f"🔍 您可以使用Excel的筛选功能来进一步分析数据")
    else:
        print(f"\n❌ 转换失败！")

if __name__ == "__main__":
    main()