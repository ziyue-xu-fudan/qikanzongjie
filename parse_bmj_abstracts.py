#!/usr/bin/env python3
"""
BMJ期刊摘要解析工具
将BMJ期刊摘要文本文件解析成结构化的表格数据
"""

import pandas as pd
import re
from datetime import datetime
from pathlib import Path
import json

class BMJAbstractParser:
    def __init__(self):
        self.articles = []
        self.current_article = {}
        self.current_section = None
        
    def parse_bmj_abstracts(self, file_path):
        """解析BMJ摘要文本文件"""
        print(f"📄 开始解析BMJ摘要文件: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 按文章分割（每个文章以数字+点号开始）
            article_pattern = r'(\d+)\.\s*BMJ\.\s*(.+?)(?=\n\n\d+\.\s*BMJ\.|\Z)'
            articles = re.findall(article_pattern, content, re.DOTALL)
            
            print(f"📊 找到 {len(articles)} 篇文章")
            
            for i, (article_num, article_content) in enumerate(articles, 1):
                print(f"🔍 解析第 {i} 篇文章...")
                article_data = self.parse_single_article(article_num, article_content)
                if article_data:
                    self.articles.append(article_data)
            
            print(f"✅ 成功解析 {len(self.articles)} 篇文章")
            return self.articles
            
        except FileNotFoundError:
            print(f"❌ 文件未找到: {file_path}")
            return []
        except Exception as e:
            print(f"❌ 解析文件时出错: {e}")
            return []
    
    def parse_single_article(self, article_num, content):
        """解析单篇文章"""
        article = {
            'article_number': int(article_num),
            'parsed_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # 提取期刊信息行
        journal_pattern = r'BMJ\.\s*(\d{4})\s+(\w+)\s+(\d+);(\d+):([e\d]+)\.\s*doi:\s*([\d\.-]+/[\w\.-]+)\.'
        journal_match = re.search(journal_pattern, content)
        
        if journal_match:
            article['journal'] = 'BMJ'
            article['pub_year'] = journal_match.group(1)
            article['pub_month'] = journal_match.group(2)
            article['pub_day'] = journal_match.group(3)
            article['volume'] = journal_match.group(4)
            article['issue'] = journal_match.group(5)
            article['doi'] = journal_match.group(6)
        
        # 提取标题
        title_pattern = r'\.\s*doi:\s*[\d\.-]+/[\w\.-]+\.\s*\n\n(.+?)\n\n'
        title_match = re.search(title_pattern, content, re.DOTALL)
        if title_match:
            title = title_match.group(1).replace('\n', ' ').strip()
            # 清理多余的空格
            title = re.sub(r'\s+', ' ', title)
            article['title'] = title
        
        # 提取作者信息
        authors_pattern = r'([A-Za-z\s\-\.]+\(\d+\)(?:,\s*[A-Za-z\s\-\.]+\(\d+\)(?:\(\d+\))*\)*)'
        authors_matches = re.findall(authors_pattern, content)
        
        if authors_matches:
            # 处理作者列表
            authors_list = []
            for author_match in authors_matches:
                # 提取作者姓名和机构编号
                author_parts = re.findall(r'([A-Za-z\s\-\.]+)\((\d+)\)', author_match)
                for name, institution_num in author_parts:
                    authors_list.append({
                        'name': name.strip(),
                        'institution_num': institution_num
                    })
            
            article['authors'] = authors_list
            article['author_count'] = len(authors_list)
            
            # 创建作者字符串
            author_names = [author['name'] for author in authors_list]
            article['authors_str'] = ', '.join(author_names)
        
        # 提取机构信息
        institution_pattern = r'\(\d+\)([A-Za-z\s,\.\-@]+?)(?=\(\d+\)|Author information:|OBJECTIVE:|CONCLUSIONS:|DOI:|Conflict|$)'
        institution_matches = re.findall(institution_pattern, content, re.DOTALL)
        
        institutions = []
        for inst in institution_matches:
            inst_clean = inst.strip()
            if inst_clean and len(inst_clean) > 10:  # 过滤掉太短的机构信息
                # 清理机构信息
                inst_clean = re.sub(r'\s+', ' ', inst_clean)
                institutions.append(inst_clean)
        
        article['institutions'] = institutions
        article['institution_count'] = len(institutions)
        
        # 提取研究目的/背景
        objective_pattern = r'OBJECTIVE:\s*(.+?)(?=DESIGN:|SETTING:|PARTICIPANTS:|MAIN|RESULTS:|CONCLUSIONS:|METHODS:|©)'
        objective_match = re.search(objective_pattern, content, re.IGNORECASE | re.DOTALL)
        if objective_match:
            objective = objective_match.group(1).strip()
            objective = re.sub(r'\s+', ' ', objective)
            article['objective'] = objective
        
        # 提取研究设计
        design_pattern = r'DESIGN:\s*(.+?)(?=SETTING:|PARTICIPANTS:|MAIN|RESULTS:|CONCLUSIONS:|METHODS:|©)'
        design_match = re.search(design_pattern, content, re.IGNORECASE | re.DOTALL)
        if design_match:
            design = design_match.group(1).strip()
            design = re.sub(r'\s+', ' ', design)
            article['design'] = design
        
        # 提取研究设置
        setting_pattern = r'SETTING:\s*(.+?)(?=PARTICIPANTS:|MAIN|RESULTS:|CONCLUSIONS:|METHODS:|©)'
        setting_match = re.search(setting_pattern, content, re.IGNORECASE | re.DOTALL)
        if setting_match:
            setting = setting_match.group(1).strip()
            setting = re.sub(r'\s+', ' ', setting)
            article['setting'] = setting
        
        # 提取参与者信息
        participants_pattern = r'PARTICIPANTS:\s*(.+?)(?=MAIN|RESULTS:|CONCLUSIONS:|METHODS:|©)'
        participants_match = re.search(participants_pattern, content, re.IGNORECASE | re.DOTALL)
        if participants_match:
            participants = participants_match.group(1).strip()
            participants = re.sub(r'\s+', ' ', participants)
            article['participants'] = participants
        
        # 提取主要结果测量
        main_outcome_pattern = r'MAIN OUTCOME MEASURES?:\s*(.+?)(?=RESULTS:|CONCLUSIONS:|METHODS:|©)'
        main_outcome_match = re.search(main_outcome_pattern, content, re.IGNORECASE | re.DOTALL)
        if main_outcome_match:
            main_outcome = main_outcome_match.group(1).strip()
            main_outcome = re.sub(r'\s+', ' ', main_outcome)
            article['main_outcome_measures'] = main_outcome
        
        # 提取结果
        results_pattern = r'RESULTS:\s*(.+?)(?=CONCLUSIONS?:|CONCLUSION:|©)'
        results_match = re.search(results_pattern, content, re.IGNORECASE | re.DOTALL)
        if results_match:
            results = results_match.group(1).strip()
            results = re.sub(r'\s+', ' ', results)
            article['results'] = results
        
        # 提取结论
        conclusions_pattern = r'CONCLUSIONS?:\s*(.+?)(?=©|Conflict|$)'
        conclusions_match = re.search(conclusions_pattern, content, re.IGNORECASE | re.DOTALL)
        if conclusions_match:
            conclusions = conclusions_match.group(1).strip()
            conclusions = re.sub(r'\s+', ' ', conclusions)
            article['conclusions'] = conclusions
        
        # 提取PMCID
        pmcid_pattern = r'PMCID:\s*(PMC\d+)'
        pmcid_match = re.search(pmcid_pattern, content)
        if pmcid_match:
            article['pmcid'] = pmcid_match.group(1)
        
        # 提取PMID
        pmid_pattern = r'PMID:\s*(\d+)\s*\[Indexed for MEDLINE\]'
        pmid_match = re.search(pmid_pattern, content)
        if pmid_match:
            article['pmid'] = pmid_match.group(1)
        
        # 提取利益冲突声明
        conflict_pattern = r'Conflict of interest statement:\s*(.+?)(?=\n\n|$)'
        conflict_match = re.search(conflict_pattern, content, re.IGNORECASE | re.DOTALL)
        if conflict_match:
            conflict = conflict_match.group(1).strip()
            conflict = re.sub(r'\s+', ' ', conflict)
            article['conflict_of_interest'] = conflict
        
        return article
    
    def create_dataframe(self):
        """创建DataFrame用于表格展示"""
        if not self.articles:
            return pd.DataFrame()
        
        # 准备数据
        data = []
        for article in self.articles:
            row = {
                '文章编号': article.get('article_number', ''),
                'PMID': article.get('pmid', ''),
                'PMCID': article.get('pmcid', ''),
                'DOI': article.get('doi', ''),
                '标题': article.get('title', ''),
                '作者': article.get('authors_str', ''),
                '作者数量': article.get('author_count', 0),
                '发表年份': article.get('pub_year', ''),
                '发表月份': article.get('pub_month', ''),
                '卷': article.get('volume', ''),
                '期': article.get('issue', ''),
                '期刊': article.get('journal', 'BMJ'),
                '研究目的': article.get('objective', ''),
                '研究设计': article.get('design', ''),
                '研究设置': article.get('setting', ''),
                '参与者': article.get('participants', ''),
                '主要结果测量': article.get('main_outcome_measures', ''),
                '结果': article.get('results', ''),
                '结论': article.get('conclusions', ''),
                '利益冲突': article.get('conflict_of_interest', ''),
                '机构数量': article.get('institution_count', 0),
                '机构': '; '.join(article.get('institutions', []))[:500],  # 限制长度
                '解析日期': article.get('parsed_date', '')
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        return df
    
    def save_to_excel(self, df, output_file):
        """保存到Excel文件"""
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 主数据表
                df.to_excel(writer, sheet_name='BMJ文献', index=False)
                
                # 统计表
                stats_df = self.generate_statistics(df)
                stats_df.to_excel(writer, sheet_name='统计信息', index=False)
                
                # 获取工作表进行格式化
                worksheet = writer.sheets['BMJ文献']
                self.format_excel_worksheet(worksheet, df)
            
            print(f"💾 Excel文件已保存: {output_file}")
            return True
            
        except Exception as e:
            print(f"❌ 保存Excel文件时出错: {e}")
            return False
    
    def format_excel_worksheet(self, worksheet, df):
        """格式化Excel工作表"""
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 标题行样式
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 数据行样式
            data_font = Font(size=10)
            data_alignment = Alignment(vertical="top", wrap_text=True)
            
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 应用标题行样式
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # 设置行高
                worksheet.row_dimensions[1].height = 30
            
            # 应用数据行样式
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                
                # 设置数据行高
                worksheet.row_dimensions[row_num].height = 60
            
            # 设置列宽
            column_widths = {
                'A': 8,   # 文章编号
                'B': 12,  # PMID
                'C': 12,  # PMCID
                'D': 25,  # DOI
                'E': 50,  # 标题
                'F': 40,  # 作者
                'G': 8,   # 作者数量
                'H': 8,   # 发表年份
                'I': 8,   # 发表月份
                'J': 8,   # 卷
                'K': 8,   # 期
                'L': 8,   # 期刊
                'M': 50,  # 研究目的
                'N': 25,  # 研究设计
                'O': 30,  # 研究设置
                'P': 40,  # 参与者
                'Q': 40,  # 主要结果测量
                'R': 50,  # 结果
                'S': 50,  # 结论
                'T': 30,  # 利益冲突
                'U': 8,   # 机构数量
                'V': 50,  # 机构
                'W': 15   # 解析日期
            }
            
            # 应用列宽
            for col_letter, width in column_widths.items():
                if col_letter <= get_column_letter(len(df.columns)):
                    worksheet.column_dimensions[col_letter].width = width
            
            # 添加筛选功能
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # 冻结首行
            worksheet.freeze_panes = 'A2'
            
            print("✅ Excel格式化完成")
            
        except Exception as e:
            print(f"⚠️  Excel格式化时出错: {e}")
    
    def generate_statistics(self, df):
        """生成统计信息"""
        stats_data = []
        
        # 基本统计
        stats_data.append(['总文章数', len(df)])
        stats_data.append(['有PMID的文章数', df['PMID'].notna().sum()])
        stats_data.append(['有DOI的文章数', df['DOI'].notna().sum()])
        stats_data.append(['平均作者数量', df['作者数量'].mean()])
        stats_data.append(['最多作者数量', df['作者数量'].max()])
        stats_data.append(['平均机构数量', df['机构数量'].mean()])
        
        # 按年份统计
        year_counts = df['发表年份'].value_counts().sort_index(ascending=False)
        for year, count in year_counts.head(10).items():
            stats_data.append([f'{year}年发表文章数', count])
        
        # 按研究设计统计
        design_counts = df['研究设计'].value_counts()
        for design, count in design_counts.head(10).items():
            if pd.notna(design) and design.strip():
                stats_data.append([f'{design}', count])
        
        stats_df = pd.DataFrame(stats_data, columns=['统计项目', '数值'])
        return stats_df

def main():
    """主函数"""
    print("📚 BMJ期刊摘要解析工具")
    print("=" * 60)
    
    # 输入文件路径
    input_file = "/Users/ziyuexu/Documents/trae_projects/paper1/abstract-BMJJournal-set (2).txt"
    
    # 输出文件路径
    output_file = "/Users/ziyuexu/Documents/trae_projects/paper1/bmj_articles_parsed.xlsx"
    
    print(f"📁 输入文件: {input_file}")
    print(f"📊 输出文件: {output_file}")
    
    # 创建解析器
    parser = BMJAbstractParser()
    
    # 解析文件
    articles = parser.parse_bmj_abstracts(input_file)
    
    if articles:
        # 创建DataFrame
        df = parser.create_dataframe()
        
        if not df.empty:
            print(f"\n📊 解析结果预览:")
            print(f"总文章数: {len(df)}")
            print(f"数据列: {', '.join(df.columns)}")
            
            # 显示前3篇文章的摘要信息
            print(f"\n📖 前3篇文章预览:")
            for i, (_, row) in enumerate(df.head(3).iterrows(), 1):
                print(f"\n{i}. {row['标题']}")
                print(f"   作者: {row['作者']}")
                print(f"   PMID: {row['PMID']}")
                print(f"   研究设计: {row['研究设计']}")
                print(f"   结论: {str(row['结论'])[:200]}...")
            
            # 保存到Excel
            success = parser.save_to_excel(df, output_file)
            
            if success:
                print(f"\n🎉 解析完成！")
                print(f"📄 Excel文件已生成: {output_file}")
                print(f"📈 包含 {len(df)} 篇BMJ文章的完整信息")
                print(f"💡 文件包含多个工作表：主数据表和统计信息表")
            else:
                print(f"\n❌ Excel文件生成失败")
        else:
            print(f"\n❌ 未能创建有效的数据表格")
    else:
        print(f"\n❌ 未能解析到任何文章")

if __name__ == "__main__":
    main()